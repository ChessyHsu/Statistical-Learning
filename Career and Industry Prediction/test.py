import jieba
import numpy as np
from openpyxl import load_workbook
import pickle
import keras
from openpyxl import load_workbook
from keras.models import Sequential 
from keras.layers import Dense, Dropout, Activation,LSTM
from keras.optimizers import SGD 
from keras import layers
from keras.models import Model
from keras import Input
from keras.models import load_model
if __name__ == '__main__':
	wb = load_workbook('test.xlsx')
	with open('job_tokenizer.pickle', 'rb') as handle:
		job_tokenizer = pickle.load(handle)
	with open('industry_tokenizer.pickle', 'rb') as handle:
		industry_tokenizer = pickle.load(handle)
	ws = wb['Sheet1']
	cnt = 0
	ind_model = load_model('industry20.hdf5')
	ind_model2 = load_model('industry8.hdf5')
	ind_model3 = load_model('industry14.hdf5')
	ind_model4 = load_model('industry19.hdf5')
	ind_model5 = load_model('industry15.hdf5')
	ind_model6 = load_model('industry12.hdf5')
	ind_model7 = load_model('industry7.hdf5')
	ind_model8 = load_model('industry16.hdf5')
	ind_model9 = load_model('industry18.hdf5')
	ind_model10 = load_model('industry17.hdf5')

	job_model = load_model('career11.hdf5')
	job_model2 = load_model('career8.hdf5')
	job_model3 = load_model('career9.hdf5')
	job_model4 = load_model('career12.hdf5')
	job_model5 = load_model('career6.hdf5')
	job_model6 = load_model('career10.hdf5')
	job_model7 = load_model('career7.hdf5')
	job_model8 = load_model('career3.hdf5')
	job_model9 = load_model('career2.hdf5')
	job_model10 = load_model('career5.hdf5')

	ou = open('ml_3.csv','w')
	f = open('job_real.csv','r')
	tmp = f.readlines()
	mp2 = np.zeros((200,))
	for x in tmp:
		s = x.split(',')
		#print(s)
		id1 = int(s[1].rstrip('\n'))
		id0 = int(s[0])
		print(id0,id1)
		mp2[id0] = id1
	mp1 = np.zeros((200,))
	f = open('ind_real.csv','r')
	tmp = f.readlines()
	for x in tmp:
		s = x.split(',')
		#print(s)
		id1 = int(s[1].rstrip('\n'))
		print(id1)
		# if id1>100:
		# 	continue
		id0 = int(s[0])
		mp1[id0] = id1
	print('x01,prediction',file=ou)
	for row in ws.rows:
		if cnt==0:
			cnt +=1
			continue
		ls = []
		tmp_ans = []

		for cell in row:
			ls.append(cell.value)
		ind = ls[0]
		
		tmp_list = []
		tmp_list2 = []
		cut = jieba.cut_for_search(ls[1])
		for x in range(len(ls[1])):
			tmp_list.append(ls[1][x])
		for x in range(len(ls[2])):
			tmp_list.append(ls[2][x])
		for x in range(len(ls[3])):
			tmp_list2.append(ls[3][x])
		for x in range(len(ls[4])):
			tmp_list2.append(ls[4][x])
		for x in range(len(ls[5])):
			tmp_list2.append(ls[5][x])
		st1 = industry_tokenizer.texts_to_sequences(tmp_list)
		st2 = job_tokenizer.texts_to_sequences(tmp_list2)
		inp = np.zeros((1,64))
		i  =0
		for x in st1:
			if len(x)!=0:
				inp[0,i] = x[0]
				i+=1
			else:
				i+=1
		#print(inp)
		p1 = ind_model.predict(inp)
		p12 = ind_model2.predict(inp)
		p13 = ind_model3.predict(inp)
		p14 = ind_model4.predict(inp)
		p15 = ind_model5.predict(inp)
		p16 = ind_model6.predict(inp)
		p17 = ind_model7.predict(inp)
		p18 = ind_model8.predict(inp)
		p19 = ind_model9.predict(inp)
		p10 = ind_model10.predict(inp)
		p1=p1+p12+0.8*p13+0.8*p14+0.6*p15+0.6*p16+0.4*p17+0.4*p18+0.2*p19+0.2*p10
		# +0.8*p14
		# print(p1.shape)
		# print(p12.shape)
		# print(p1)
		# print(p12)
		# input()
		res1  = np.argmax(p1)
		print(res1)
		inp = np.zeros((1,65))
		i  =0
		for x in st2:
			if len(x)!=0:
				inp[0,i] = x[0]
				i+=1
			else:
				i+=1
		#print(inp)
		p2 = job_model.predict(inp)
		p22 = job_model2.predict(inp)
		p23 = job_model3.predict(inp)
		p24 = job_model4.predict(inp)
		p25 = job_model5.predict(inp)
		p26 = job_model6.predict(inp)
		p27 = job_model7.predict(inp)
		p28 = job_model8.predict(inp)
		p29 = job_model9.predict(inp)
		p20 = job_model10.predict(inp)

		p2=p2+p22+0.8*p23+0.8*p24+0.6*p25+0.6*p26+0.4*p27+0.4*p28+0.2*p29+0.2*p20
		# +0.8*p24+0.8*p25
		# print(p2)
		res2  = np.argmax(p2)
		print(res2)
		print(str(ind)+'_a08a01,'+str(int(mp1[res1])),file=ou)
		print(str(ind)+'_a08a02,'+str(int(mp2[res2])),file=ou)
		#st2 = ls[3]
		#print(ls)