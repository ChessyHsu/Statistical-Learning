import keras
from openpyxl import load_workbook
from keras.models import Sequential 
from keras.layers import Dense, Dropout, Activation,LSTM, Bidirectional, BatchNormalization, GRU
from keras.optimizers import SGD, RMSprop
from keras import layers
from keras.models import Model
from keras import Input
import numpy as np
import jieba
import pickle
from keras.callbacks import ModelCheckpoint, EarlyStopping,  ReduceLROnPlateau
from gensim.test.utils import common_texts, get_tmpfile
from gensim.models import Word2Vec

if __name__ == '__main__':
	wb = load_workbook('train.xlsx')
	#print(wb.sheetnames)
	flag = np.zeros((3167,64))
	ws = wb['Sheet1']
	cnt = 0
	corpus = []
	f = open('ind_real.csv')
	tmp = f.readlines()
	mp = np.zeros((100))
	for x in tmp:
		s = x.split(',')
		#print(s)
		id1 = int(s[1].rstrip('\n'))
		print(id1)
		if id1>100:
			continue
		mp[id1] = int(s[0])
	#tokenizer.fit_on_texts(corpus)
	for row in ws.rows:
		if cnt==0:
			cnt +=1
			continue
		
		#print(cnt)
		ls = []
		tmp_ans = []
		for cell in row:
			ls.append(cell.value)
		#com = ls[3]+ls[4]+ls[5]
		ind = int(ls[1])
		if ind > 100:
			continue
		ind = int(mp[ind])
		flag[cnt-1][ind] = 1
		#cut = jieba.cut_for_search(ls[3])
		tmp_list = []
		for x in range(len(ls[3])):
			tmp_list.append(ls[3][x])
		for x in range(len(ls[4])):
			tmp_list.append(ls[4][x])
		# cut = jieba.cut_for_search(ls[5])
		# for x in cut:
		# 	tmp_list.append(x)
		#print(tmp_list)
		corpus.append(tmp_list)
		cnt+=1
	MAX_NUM_WORDS = 100
	EMB_DIM = 256

	##train embedding
	wmodel = Word2Vec(corpus,workers=10, min_count=1, size=EMB_DIM)
	wmodel.train(corpus, total_examples=wmodel.corpus_count, epochs=1000)
	wmodel.save("word2vec.model")

	#print(flag[0])

	word2index = {}
	for word in wmodel.wv.vocab:
		word2index[word] = len(word2index)

	word2index['PAD'] = len(word2index)
	word2index['UNK'] = len(word2index)

	x1_train = []
	for cnt, sen in enumerate(corpus):
		x1_train.append([])
		for word in sen:
			if word in word2index.keys():
				x1_train[cnt].append(word2index[word])
			else:
				x1_train[cnt].append(word2index["UNK"])

	for sen in x1_train:
		if len(sen) < MAX_NUM_WORDS:
			sen += [word2index["PAD"]] * (MAX_NUM_WORDS-len(sen))
		else:
			sen = sen[:MAX_NUM_WORDS]

	num_words = len(wmodel.wv.vocab) + 2
	emb_weight = np.zeros((num_words, EMB_DIM), dtype=float)
	for cnt, word in enumerate(wmodel.wv.vocab):
		emb_weight[cnt] = wmodel.wv[word]

	vector = np.random.uniform(1, EMB_DIM)
	emb_weight[num_words-1] = vector
	vector = np.random.uniform(1, EMB_DIM)
	emb_weight[num_words-2] = vector
	x1_train = np.array(x1_train)
	print(x1_train.shape)
	# #print(len(corpus))
	# x1_train = []
	# for x in range(3167):
	# 	x1_train.append(tokenizer.texts_to_sequences(corpus[x]))
################
	# max_seq_len = max([len(seq) for seq in x1_train])
	# #print(max_seq_len)
	# x_train = np.zeros((3167,40))
	# for x in range(3167):
	# 	for y in range(len(x1_train[x])):
	# 		#print(x1_train[x][y])
	# 		ls = x1_train[x][y]
	# 		if len(ls)==0:
	# 			continue
	# 		x_train[x][y] = ls[0]
	
	#print(x_train[0])
		#print(len(x1_train[x]))
	#print(x1_train)
	#print(x1_train[0])
		#model.add(embedding_layer)
	#embedding_layer = layers.Embedding(MAX_NUM_WORDS, 256)
	# NUM_LSTM_UNITS = 256
	#model = Sequential()
	#op_input = Input(shape=(25, ), dtype='int32')

	# 詞嵌入層
	# 經過詞嵌入層的轉換，兩個新聞標題都變成
	# 一個詞向量的序列，而每個詞向量的維度
	# 為 256
	#embedding_layer = Embedding(MAX_NUM_WORDS, NUM_EMBEDDING_DIM)
	model = Sequential()
	model.add(layers.Embedding(num_words,EMB_DIM,weights=[emb_weight],
		input_length=MAX_NUM_WORDS, trainable=False))
	# LSTM 層
	# 兩個新聞標題經過此層後
	# 為一個 128 維度向量
	#model.add(LSTM(256,return_sequences=True))
	#model.add(Activation('tanh'))s
	#model.add(Dropout(0.25))
	#model.add(LSTM(256,return_sequences=True))
	#model.add(Activation('tanh'))
	
	#model.add(LSTM(NUM_LSTM_UNITS,return_sequences=True))
	#model.add(Activation('tanh'))
	# model.add(LSTM(256, dropout=0.1, recurrent_dropout=0.5, return_sequences=True))
	model.add(Bidirectional(GRU(256, dropout=0.1, recurrent_dropout=0.7, return_sequences=True)))
	model.add(BatchNormalization())
	model.add(Bidirectional(GRU(256, dropout=0.1, recurrent_dropout=0.7, return_sequences=False)))
	model.add(BatchNormalization())
	# model.add(Activation('tanh'))
	#model.add(Dropout(0.25))
	model.add(Dense(units=64, activation='softmax'))
	#output = shared_lstm(embedded)
	#output = t1(output)
	#output = shared_lstm(output)
	#output = t1(output)
	#bm_output = shared_lstm(bm_embedded)

	# 串接層將兩個新聞標題的結果串接單一向量
	# 方便跟全連結層相連
	#merged = concatenate([top_output, bm_output], axis=-1)

# 全連接層搭配 Softmax Activation
# 可以回傳 3 個成對標題
# 屬於各類別的可能機率  
	#predictions = dense(output)
	# with open('industry_tokenizer.pickle', 'wb') as handle:
 # 		pickle.dump(tokenizer, handle, protocol=pickle.HIGHEST_PROTOCOL)
	#predictions = dr(predictions)
	#model = Model(inputs=top_input, outputs=predictions)
	# rms = RMSprop(lr=0.001, rho=0.9, epsilon=None, decay=0.001)
	rms=keras.optimizers.RMSprop(lr=0.01)
	reduce_lr = ReduceLROnPlateau(monitor='val_loss', factor=0.2,
                                patience=1,verbose=0,
                                mode='auto', min_delta=0.0001,
                                cooldown=0, min_lr=0.001)
	model.compile(optimizer=rms,loss='categorical_crossentropy',metrics=['accuracy'])
	earlystopping = EarlyStopping(monitor='val_acc', patience = 30, verbose=1, mode='max') 
	checkpoint = ModelCheckpoint(filepath='industry7.hdf5', verbose=1, save_best_only=True, \
								save_weights_only=False, monitor='val_acc', mode='max')
	#sgd = SGD(lr=0.01, decay=1e-6, momentum=0.9, nesterov=True)
	model.summary()
	model.fit(x1_train,flag, batch_size=512,epochs=300,shuffle=True,verbose=1,
		validation_split=0.1, callbacks=[checkpoint,earlystopping, reduce_lr])

