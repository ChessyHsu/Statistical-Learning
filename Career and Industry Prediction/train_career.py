import keras
from openpyxl import load_workbook
from keras.models import Sequential 
from keras.layers import Dense, Dropout, Activation, LSTM, Bidirectional
from keras.optimizers import SGD 
from keras import layers
from keras.models import Model
from keras import Input
import numpy as np
import jieba
from keras.callbacks import ModelCheckpoint, EarlyStopping, ReduceLROnPlateau
import pickle
if __name__ == '__main__':
	wb = load_workbook('train.xlsx')
	#print(wb.sheetnames)
	flag = np.zeros((3200,120))
	ws = wb['Sheet1']
	cnt = 0
	corpus = []
	f = open('job_real.csv')
	tmp = f.readlines()
	mp = np.zeros((10000,))
	for x in tmp:
		s = x.split(',')
		#print(s)
		id1 = int(s[1].rstrip('\n'))
		# print(id1)
		# if id1>100:
		# 	continue
		mp[id1] = int(s[0])
	for row in ws.rows:
		if cnt==0:
			cnt +=1
			continue
		
		# print(cnt)
		ls = []
		tmp_ans = []
		for cell in row:
			ls.append(cell.value)
		#com = ls[3]+ls[4]+ls[5]
		#ind = ls[2]
		#print(ind)
		# if ind > 100:
		# 	continue
		tmp_list = []
		ind = int(ls[2])
		ind = int(mp[ind])
		flag[cnt-1][ind] = 1
		for x in range(len(ls[5])):
			tmp_list.append(ls[5][x])
		for x in range(len(ls[6])):
			tmp_list.append(ls[6][x])
		for x in range(len(ls[7])):
			tmp_list.append(ls[7][x])
		# cut = jieba.cut_for_search(ls[5])
		# for x in cut:
		# 	tmp_list.append(x)
		# print(tmp_list)
		#print('hi')
		corpus.append(tmp_list)
		cnt+=1
	MAX_NUM_WORDS = 3000

	#print(flag[0])
	tokenizer = keras.preprocessing.text.Tokenizer(num_words=MAX_NUM_WORDS)
	tokenizer.fit_on_texts(corpus)
	#print(len(corpus))
	#input()
	# print(len(corpus))
	x1_train = []
	for x in range(3200):
		x1_train.append(tokenizer.texts_to_sequences(corpus[x]))

	
	max_seq_len = max([len(seq) for seq in x1_train])
	# print(max_seq_len)
	x_train = np.zeros((3200,65))
	for x in range(3200):
		for y in range(len(x1_train[x])):
			#print(x1_train[x][y])
			ls = x1_train[x][y]
			if len(ls)==0:
				continue
			x_train[x][y] = ls[0]
	
	# print(x_train[0])
		#print(len(x1_train[x]))
	#print(x1_train)
	#print(x1_train[0])
		#model.add(embedding_layer)
	#embedding_layer = layers.Embedding(MAX_NUM_WORDS, 256)
	# NUM_LSTM_UNITS = 256
	#model = Sequential()
	#op_input = Input(shape=(25, ), dtype='int32')

	# 詞嵌入層
	# 經過詞嵌入層的轉換，兩個新聞標題都變成
	# 一個詞向量的序列，而每個詞向量的維度
	# 為 256
	#embedding_layer = Embedding(MAX_NUM_WORDS, NUM_EMBEDDING_DIM)
	model = Sequential()
	model.add(layers.Embedding(MAX_NUM_WORDS, 1024))
	# LSTM 層
	# 兩個新聞標題經過此層後
	# 為一個 128 維度向量
	#model.add(LSTM(256,return_sequences=True))
	#model.add(Activation('tanh'))
	#model.add(Dropout(0.25))
	#model.add(LSTM(256,return_sequences=True))
	#model.add(Activation('tanh'))
	
	#model.add(LSTM(NUM_LSTM_UNITS,return_sequences=True))
	#model.add(Activation('tanh'))
	#model.add(LSTM(64))
	#model.add(Activation('tanh'))
	#model.add(Dropout(0.25))
	# model.add(LSTM(units=512))
	# model.add(LSTM(256, dropout=0.1, recurrent_dropout=0.9))
	model.add(Bidirectional(LSTM(256, dropout=0.1, recurrent_dropout=0.5, return_sequences=True)))
	model.add(Bidirectional(LSTM(256, dropout=0.1, recurrent_dropout=0.9, return_sequences=False)))
	# model.add(Dropout(0.25))
	# model.add(Activation('tanh'))
	#model.add(Dense(units=1000))
	# model.add(Dropout(0.25))
	model.add(Dense(units=120, activation='softmax'))
	#output = shared_lstm(embedded)
	#output = t1(output)
	#output = shared_lstm(output)
	#output = t1(output)
	#bm_output = shared_lstm(bm_embedded)

	# 串接層將兩個新聞標題的結果串接單一向量
	# 方便跟全連結層相連
	#merged = concatenate([top_output, bm_output], axis=-1)

# 全連接層搭配 Softmax Activation
# 可以回傳 3 個成對標題
# 屬於各類別的可能機率  
	#predictions = dense(output)
	
	#predictions = dr(predictions)
	#model = Model(inputs=top_input, outputs=predictions)
	with open('job_tokenizer.pickle', 'wb') as handle:
 		pickle.dump(tokenizer, handle, protocol=pickle.HIGHEST_PROTOCOL)
	rms=keras.optimizers.RMSprop(lr=0.001)
	reduce_lr = ReduceLROnPlateau(monitor='val_loss', factor=0.2,
                                patience=1,verbose=0,
                                mode='auto', min_delta=0.0001,
                                cooldown=0, min_lr=0.001)
	sgd = SGD(lr=0.01, decay=1e-6, momentum=0.9, nesterov=True)
	model.compile(optimizer=rms ,loss='categorical_crossentropy',metrics=['accuracy'])
	model.summary()
	earlystopping = EarlyStopping(monitor='val_acc', patience = 30, verbose=1, mode='max') 
	checkpoint = ModelCheckpoint(filepath='career13.hdf5', verbose=1, save_best_only=True, \
								save_weights_only=False, monitor='val_acc', mode='max')
	#sgd = SGD(lr=0.01, decay=1e-6, momentum=0.9, nesterov=True)
	model.summary()
	model.fit(x_train,flag,batch_size=128,epochs=300,shuffle=True,verbose=1,
		validation_split=0.1, callbacks=[checkpoint,earlystopping,reduce_lr])

